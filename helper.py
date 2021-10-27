import locale
import re
from userController import input_dates
from openpyxl import styles
import os

# Check String caracter when user entered string caracter
def check_tanggal(dates, total_days):
    sort_dates = dates.split('-')
    dates_chooses = dates.split(',')

    check_string_caracter = re.findall('[a-zA-Z]', dates)


    if len(check_string_caracter) > 0:
        return input_dates('Masukan angka jangan string')
    elif '' in sort_dates or '' in dates_chooses:
        return input_dates('Masukan tanggal sesuai intruksi')
    elif(len(sort_dates) == 2 and '-' in dates):

        sort_dates = [int(i) for i in sort_dates]
        if sort_dates[0] < sort_dates[1] and sort_dates[1] < total_days:
            return sort_dates, '-'
        else:
            return input_dates('Maaf tanggal tidak boleh kebalik atau lebih dari dates bulan ini')

    elif len(dates_chooses) > 1 and ',' in dates:

        due_date = [int(i) < total_days for i in dates_chooses]
        if all(due_date):
            dates_chooses = [int(i) for i in dates_chooses]
            filter_same_number = set(dates_chooses)
            print('tanggal yang sama akan mengambil 1 datesan yang sama dari beberapa tanggal contoh :', filter_same_number)
            return list(filter_same_number), ','
        else:
            return input_dates('Maaf anda melebihi batas tanggal bulan ini')
    elif int(dates) < total_days:
        return [int(dates)], ','
    else:
        return input_dates('Maaf masukan angka yang valid')
        

def format_rupiah(duit):
    locale.setlocale(locale.LC_NUMERIC, 'IND')
    rupiah = locale.format("%.*f", (2, duit), True)
    return "Rp " + rupiah



def store_data_omset(dates, ws, total_money_omset, clock_input, separator):
    max_row = ws.max_row + 1
    i = 0
    if len(dates) == 2 and '-' in separator:
        for date in range(dates[0], dates[1] + 1):
            ws["A{}".format(max_row + i)].alignment = styles.Alignment(horizontal='left', vertical='center')
            ws["B{}".format(max_row + i)].alignment = styles.Alignment(horizontal='left', vertical='center')
            ws["C{}".format(max_row + i)].alignment = styles.Alignment(horizontal='left', vertical='center')
            
            ws['A{}'.format(max_row + i)] = date
            ws['B{}'.format(max_row + i)] = total_money_omset[i]
            ws['C{}'.format(max_row + i)] = clock_input
            i = i + 1
    elif len(dates) > 0 and ',' in separator:
        for date in dates:
            ws["A{}".format(max_row + i)].alignment = styles.Alignment(horizontal='left', vertical='center')
            ws["B{}".format(max_row + i)].alignment = styles.Alignment(horizontal='left', vertical='center')
            ws["C{}".format(max_row + i)].alignment = styles.Alignment(horizontal='left', vertical='center')
            
            ws['A{}'.format(max_row + i)] = date
            ws['B{}'.format(max_row + i)] = total_money_omset[i]
            ws['C{}'.format(max_row + i)] = clock_input
            i = i + 1
    else:
        print('Eh ada yang error gaes nanti aja')

    return ws


def create_directory_check(file):
    if os.path.isdir('excel'):
        if os.path.isfile(f'./excel/{file}.xlsx'):
            return True, file
        else:
            return False, file
    else:
        os.makedirs('excel')
        return False, file

