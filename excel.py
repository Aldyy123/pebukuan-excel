from openpyxl import Workbook, styles, load_workbook, worksheet
from openpyxl.utils.exceptions import SheetTitleException
from colorama import Fore
import re
from helper import store_data_omset
from datetime import datetime

# Memasukan hasil excel untuk di olah datanya kemudian
# save data
# memasukan file excel yang sudah ada
# update data excel


def input_new_excel(name_file, dates, dates_now, omset, separator):
    wb = Workbook()
    ws = wb.active
    ws.title = datetime.now().strftime("%B")


    ws['A1'] = 'Tanggal'
    ws['B1'] = 'Omset'
    ws['C1'] = 'Waktu'

    store_data_omset(dates, ws, omset, dates_now, separator)

    print("Berhasil di input")
    
    return wb.save('./excel/{}.xlsx'.format(name_file))


def update_excel(name_file, dates, clock_input, total_money_omset, separator):
    try:

        wb = load_workbook(f'./excel/{name_file}.xlsx')
        qa_new_sheet = input("Apakah ingin membuat sheet baru? : ")
        ws = None
        if qa_new_sheet in ['y', 'iya', 'yes', 'ya']:
            qa_new_sheet = input("Masukan nama sheet yang anda inginkan : ")
            ws = wb.create_sheet(qa_new_sheet)
        else:
            chose_sheet = input("Masukan sheet yang ingin anda ubah : ")
            ws = wb[chose_sheet]        
    except Exception:
        print("Maaf ada masalah saat pemilihan sheet excel")
        return update_excel(name_file, dates, clock_input, total_money_omset, separator)

    store_data_omset(dates, ws, total_money_omset, clock_input, separator)
    print("Berhasil di ubah")
    wb.save('./excel/{}.xlsx'.format(name_file))


def user_omset_daily(dates_omset_pengguna, clock_input, alert=''):
    input_omset_user = input("""
    \t  Format jika anda memasukan lebih dari jumlah omset harian adalah
    \t  Example : {0} 2000, 90000, 10000 {1}
    \t  Sesuai jumlah tanggal yang anda input
    \t  Anda harus menginput sesuai dengan tanggal omset yang telah di inputkan seperti yang bisa anda lihat disini {0}{4}{1}
    \t  Silahkan masukan omset hari ini: 
    \t  {3}{2}{1}
""".format(Fore.GREEN, Fore.RESET, alert, Fore.RED, dates_omset_pengguna))
    single_omset = input_omset_user
    check_string_caracter = re.findall('[a-zA-Z]', input_omset_user)
    input_omset_user = input_omset_user.split(',')

    if len(check_string_caracter) > 0 or input_omset_user[0] == '':
        return user_omset_daily(
                dates_omset_pengguna, clock_input, 'Maaf Bro harus angka tidak boleh teks')
    elif len(dates_omset_pengguna[0]) == 2  and '-' in dates_omset_pengguna[1]:
        jumlah_hari = dates_omset_pengguna[0][1] - dates_omset_pengguna[0][0]
        if len(input_omset_user) == jumlah_hari + 1:
            return [int(i) for i in input_omset_user]
        else:
            return user_omset_daily(
                dates_omset_pengguna, clock_input, 'Maaf input omset anda kurang atau kelebihan, silahkan lihat contoh format')
    elif len(dates_omset_pengguna[0]) > 1 and ',' in dates_omset_pengguna[1]:
        if len(dates_omset_pengguna[0]) == len(input_omset_user):
            return [int(i) for i in input_omset_user]
        else:
            return user_omset_daily(
                dates_omset_pengguna, clock_input, 'Maaf input omset anda kurang atau kelebihan, silahkan lihat contoh format')
    else:
        if ',' in single_omset:
            return user_omset_daily(
                dates_omset_pengguna, clock_input, 'Maaf ini input tunggal, tidak boleh ada koma')
        else:
            return [single_omset]


