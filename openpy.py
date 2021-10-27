import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from matplotlib import style
from openpyxl import load_workbook
import os



apa = pd.read_excel('./excel/iya.xlsx')

print(apa['Tanggal'].values)

x = []
label = ()
for i in range(apa.index.start, apa.index.stop):
    x.append(i)


print(x)
# book = load_workbook('./excel/iya.xlsx')
# sheet = book['October']

# a1 = sheet['B4']
# a2 = sheet['A4']
# a3 = sheet.cell(row=3, column=1)
# apa = sheet.iter_cols(2)
# print(apa)

# book.save('baru.xlsx')

# dfs = pd.read_excel('baru.xlsx', sheet_name='Oktober')

# print(dfs)

# style.use('ggplot')

# x = [0, 1, 2, 3, 4, 8]
# y = [24.27, 23.18, 22.39, 8.41, 7.19, 6.62]
y = apa['Omset'].values
print(y)
# fig, ax = plt.subplots()

# ax.bar(x, y, align='center')

# ax.set_title('Omset bulan Oktober')
# ax.set_ylabel(apa.columns[1])
# ax.set_xlabel(apa.columns[0])

# ax.set_xticks(x)
# ax.set_xticklabels(apa['Tanggal'].values)

# plt.show()
